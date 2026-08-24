#!/usr/bin/env python3
"""Curación de una vez: fondo padre de CAFCI -> id interno de la CNV, para linkear desde la ficha
de un FCI a la "COMPOSICIÓN DE CARTERA" pública que la CNV publica semanalmente (artículo 34).

Puente para el enlace CNV en la ficha de FCI (F-057/F-046). La página destino funciona sin query
params — `https://www.cnv.gov.ar/SitioWeb/FondosComunesInversion/DetallesFCI/{id}` (verificado en
vivo el 23/08/2026, HTTP 200 con sección "COMPOSICIÓN DE CARTERA")— pero el `{id}` es un
identificador interno de la CNV que ninguna fuente que ya consumimos declara: ni la planilla de
CAFCI (que trae `codigo_cnv`, el número de *registro*, no este id) ni `condiciones_emision.csv`.
Se resuelve una sola vez, a mano y verificado, y no en cada request (regla 11 del dominio).

**Fuente del listado de la CNV**: `POST /SitioWeb/FondosComunesInversion/GetFCIPorTipo` con
`tipo=1` — devuelve un único JSON con los 1.677 fondos activos (`{"Text": " NOMBRE", "Value": "id"}`
por fondo), sin paginar. Se probó también el listado HTML paginado
(`POST .../Busqueda?pagina=N`, ~56 páginas) que trae denominación, gerente y depositaria además del
id, pero el JSON alcanza para lo que necesita esta curación (matching sólo por nombre) y es un
único pedido en vez de 56: se usa el JSON.

**Fuente de los fondos padre**: la planilla diaria de CAFCI (mismo XLSX que consume
`app/ingesta/cafci/parser.py`), agrupada por `codigo_cnv` (columna S) — un fondo padre puede tener
varias clases (columna A: "Fondo X - Clase A", "Fondo X - Clase B"...) y todas comparten
`codigo_cnv`.

**Matching**: nombre normalizado (NFD sin diacríticos, minúsculas, sin "- Clase X" ni sufijos
legales de forma societaria del fondo — "F.C.I.", "FCI", "Fondo Común de Inversión (Abierto)" y
variantes cerradas—, sólo [a-z0-9]). Conservador: sólo entra a `data/fci_cnv_ids.csv` un
`codigo_cnv` cuyo nombre normalizado matchea un único id en el listado de la CNV. Cero matches o
más de uno van a `data/fci_cnv_pendientes.csv` para revisión humana — nunca se elige "el más
parecido" (mismo criterio que `tools/curar_emisores_cuit.py`).

**Dos deformaciones del listado de la CNV, medidas el 24/08/2026**, que la primera curación no
contemplaba y dejaban afuera a 123 fondos padre que sí están publicados:

1. **El nombre anterior pegado con "Ex"**: a los fondos renombrados la CNV les concatena su
   denominación vieja — `Balanz Institucional` figura como *"Balanz Institucional Ex Balanz Capital
   Multimercado I"*. Son 137 fondos. Se indexan **las dos formas** (completa y hasta antes del
   "Ex"), no se reemplaza una por otra: si un fondo se llamara legítimamente "… Ex …", su nombre
   entero sigue siendo buscable.
2. **Nombres truncados a 50 caracteres**: 125 entradas vienen cortadas a mitad de palabra
   (*"Balanz Acciones Ex Balanz Capital Acciones Argenti"*). Para ésas —y sólo para ésas, las de
   largo exactamente 50— se acepta que la denominación de la CNV sea **prefijo** del nombre de
   CAFCI, con dos recaudos contra el falso positivo: el prefijo normalizado debe medir al menos
   `LARGO_MINIMO_PREFIJO` caracteres, y si empata con más de un id la fila va a pendientes (sin
   esto, un truncado "Balanz Performance I" se comería a "Balanz Performance III").

Además se valida que **ningún id de la CNV quede asignado a dos `codigo_cnv` distintos**: si dos
fondos padre resuelven al mismo detalle, los dos van a pendientes en vez de publicar un enlace que
apunta al fondo equivocado. El CSV declara en `verificado_por` por cuál de las tres reglas entró
cada fila, para que la revisión humana pueda priorizar las menos directas.

Uso (con acceso a red, sin backend local necesario — las dos fuentes son públicas):
    python3 tools/curar_fci_cnv.py --dry-run
    python3 tools/curar_fci_cnv.py
"""

import argparse
import csv
import io
import os
import re
import sys
import unicodedata
from datetime import date

import httpx
import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
FCI_CNV_CSV = os.path.join(BASE_DIR, "data", "fci_cnv_ids.csv")
PENDIENTES_CSV = os.path.join(BASE_DIR, "data", "fci_cnv_pendientes.csv")

CNV_LISTADO_URL = "https://www.cnv.gov.ar/SitioWeb/FondosComunesInversion/GetFCIPorTipo"
CAFCI_URL = "https://api.pub.cafci.org.ar/pb_get"
USER_AGENT = "10-Swaper (asesor ALyC Argentina) moyanojjeronimo@gmail.com"

# Columnas 1-indexadas de la planilla de CAFCI, tal como las declara
# `app/ingesta/cafci/parser.py` (verificadas de nuevo contra el archivo real el 23/08/2026).
FILA_INICIO_DATOS = 10
COL_FONDO = 1
COL_MONEDA = 2
COL_CODIGO_CNV = 19
COL_GERENTE = 24

# Sufijos de forma societaria/estado del fondo que las dos fuentes escriben distinto y que no
# aportan nada para identificar el fondo. Se sacan sólo para comparar, nunca para lo que se
# guarda en el CSV final (`denominacion_cnv` y el nombre de CAFCI viajan tal como los trae cada
# fuente).
_PATRON_CLASE = re.compile(r"-?\s*clase\s+[a-z0-9ivx]+\b")
_PATRON_ESTADO = re.compile(
    r"\(\s*(en proceso de liquidaci[oó]n|valor final de liquidaci[oó]n)\s*\)"
)
# La CNV le pega a cada fondo renombrado su denominación anterior: "Nuevo Nombre Ex Nombre Viejo".
_PATRON_EX = re.compile(r"\s+ex\s+", re.IGNORECASE)

# Largo al que el listado de la CNV corta las denominaciones largas (medido: 125 entradas de 1.677
# miden exactamente esto, varias partidas a mitad de palabra).
LARGO_TRUNCADO = 50

# Piso para aceptar un match por prefijo. Un prefijo corto ("balanz") matchearía media gestora;
# con este piso, lo que se compara ya es un nombre de fondo y no una marca.
LARGO_MINIMO_PREFIJO = 25

_PATRONES_FORMA_LEGAL = (
    re.compile(r"\bfondo\s+comun\s+de\s+inversion(?:es)?\s+cerrado\s+agropecuario\b"),
    re.compile(r"\bfondo\s+comun\s+de\s+inversion(?:es)?\s+cerrado\s+inmobiliario\b"),
    re.compile(r"\bfondo\s+comun\s+de\s+inversion(?:es)?\s+abierto\b"),
    re.compile(r"\bfondo\s+comun\s+de\s+inversion(?:es)?\b"),
    # La CNV corta a 50 caracteres y parte el sufijo legal a la mitad ("… Fondo Comun de Inver").
    # Sólo al final del texto: en el medio, un corte así no existiría.
    re.compile(r"\bfondo\s+comun\s+de\s+inv[a-z]*\s*$"),
    re.compile(r"\bf\.?\s*c\.?\s*i\.?\b"),
    # Designación legal de los FCI de infraestructura (RG 900): la llevan todos los de esa clase,
    # así que no distingue un fondo de otro. Se acepta truncada por la misma razón que arriba.
    re.compile(r"\bpara\s+el\s+financiamiento\s+de\s+la\s+infraestructura\s+y\s+la\s+economia\s+rea[l]?\s*$"),
)


def normalizar(nombre: str) -> str:
    """NFD sin diacríticos, minúsculas, sin clase ni forma legal del fondo, sólo [a-z0-9]."""
    sin_tildes = unicodedata.normalize("NFD", nombre)
    sin_tildes = "".join(c for c in sin_tildes if unicodedata.category(c) != "Mn")
    s = sin_tildes.lower()
    s = _PATRON_ESTADO.sub(" ", s)
    s = _PATRON_CLASE.sub(" ", s)
    for patron in _PATRONES_FORMA_LEGAL:
        s = patron.sub(" ", s)
    return re.sub(r"[^a-z0-9]", "", s)


def variantes_normalizadas(denominacion: str) -> set[str]:
    """Las formas bajo las que se indexa una denominación de la CNV: la completa y, si arrastra el
    nombre anterior con "Ex", también la parte previa. Las dos, nunca una en lugar de la otra."""
    variantes = {normalizar(denominacion)}
    antes_del_ex = _PATRON_EX.split(denominacion, maxsplit=1)[0]
    if antes_del_ex != denominacion:
        variantes.add(normalizar(antes_del_ex))
    return {v for v in variantes if v}


def listado_cnv(cliente: httpx.Client) -> list[tuple[str, str]]:
    """`[(id, denominacion), ...]` de los 1.677 fondos activos que declara la CNV."""
    respuesta = cliente.post(
        CNV_LISTADO_URL,
        data={"tipo": "1"},
        headers={"User-Agent": USER_AGENT},
    )
    respuesta.raise_for_status()
    datos = respuesta.json()
    return [(str(item["Value"]), str(item["Text"]).strip()) for item in datos]


def fondos_padre_cafci(cliente: httpx.Client) -> dict[str, tuple[str, str | None]]:
    """`{codigo_cnv: (nombre_representativo, gerente)}`, un fondo padre por `codigo_cnv` — la
    planilla trae una fila por clase y todas comparten `codigo_cnv`."""
    respuesta = cliente.get(CAFCI_URL, headers={"User-Agent": USER_AGENT})
    respuesta.raise_for_status()
    libro = openpyxl.load_workbook(io.BytesIO(respuesta.content), data_only=True)
    ws = libro.worksheets[0]

    padres: dict[str, tuple[str, str | None]] = {}
    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):
        fondo = ws.cell(row=fila, column=COL_FONDO).value
        if fondo is None:
            continue
        moneda = ws.cell(row=fila, column=COL_MONEDA).value
        if moneda is None:
            continue  # fila de sección, no de datos
        codigo_cnv = ws.cell(row=fila, column=COL_CODIGO_CNV).value
        if codigo_cnv is None:
            continue
        codigo_cnv = str(codigo_cnv).strip()
        if codigo_cnv in padres:
            continue
        gerente = ws.cell(row=fila, column=COL_GERENTE).value
        padres[codigo_cnv] = (str(fondo).strip(), str(gerente).strip() if gerente else None)
    return padres


def main() -> None:
    ap = argparse.ArgumentParser(description="Curación fondo padre CAFCI -> id interno CNV")
    ap.add_argument("--dry-run", action="store_true", help="Muestra el resultado sin escribir nada")
    args = ap.parse_args()

    with httpx.Client(timeout=30.0, follow_redirects=True) as cliente:
        print("Descargando listado de fondos de la CNV...")
        cnv = listado_cnv(cliente)
        print(f"  {len(cnv)} fondos en el listado de la CNV.\n")

        print("Descargando planilla diaria de CAFCI...")
        padres = fondos_padre_cafci(cliente)
        print(f"  {len(padres)} fondos padre distintos (por codigo_cnv) en la planilla.\n")

    indice_cnv: dict[str, list[tuple[str, str]]] = {}
    for id_cnv, denominacion in cnv:
        for variante in variantes_normalizadas(denominacion):
            indice_cnv.setdefault(variante, []).append((id_cnv, denominacion))

    # Sólo las denominaciones que la CNV cortó: son las únicas habilitadas a matchear por prefijo.
    truncadas = [
        (normalizar(d), id_cnv, d) for id_cnv, d in cnv if len(d) == LARGO_TRUNCADO
    ]
    print(f"  {len(truncadas)} denominaciones truncadas a {LARGO_TRUNCADO} caracteres.\n")

    confirmados: list[tuple[str, str, str, str]] = []  # (codigo_cnv, id, denominacion, verificado_por)
    pendientes: list[tuple[str, str, str]] = []  # (codigo_cnv, nombre_cafci, candidatos)

    for codigo_cnv, (nombre_cafci, _gerente) in sorted(padres.items(), key=lambda kv: kv[0]):
        objetivo = normalizar(nombre_cafci)
        candidatos = indice_cnv.get(objetivo, [])
        ids_distintos = {c for c, _ in candidatos}
        regla = "nombre"

        if not ids_distintos:
            # Nada exacto: puede ser una denominación que la CNV cortó a la mitad. Se acepta que la
            # suya sea prefijo de la nuestra, nunca al revés, y con piso de largo.
            por_prefijo = [
                (id_cnv, denominacion)
                for norm_cnv, id_cnv, denominacion in truncadas
                if len(norm_cnv) >= LARGO_MINIMO_PREFIJO and objetivo.startswith(norm_cnv)
            ]
            candidatos = por_prefijo
            ids_distintos = {c for c, _ in por_prefijo}
            regla = "nombre_truncado"

        if len(ids_distintos) == 1:
            id_cnv, denominacion = candidatos[0]
            if regla == "nombre" and normalizar(denominacion) != objetivo:
                regla = "nombre_sin_ex"
            confirmados.append((codigo_cnv, id_cnv, denominacion, regla))
        elif len(ids_distintos) > 1:
            detalle = "; ".join(f"{d!r} (id {i})" for i, d in candidatos)
            pendientes.append((codigo_cnv, nombre_cafci, f"ambiguo: {detalle}"))
        else:
            pendientes.append((codigo_cnv, nombre_cafci, "sin candidatos"))

    # Un mismo detalle de la CNV no puede ser la composición de dos fondos padre distintos: si pasa,
    # al menos uno está mal apuntado y no hay forma de saber cuál. Los dos a revisión.
    por_id: dict[str, list[tuple[str, str, str, str]]] = {}
    for confirmado in confirmados:
        por_id.setdefault(confirmado[1], []).append(confirmado)

    colisiones = {id_cnv: filas for id_cnv, filas in por_id.items() if len(filas) > 1}
    if colisiones:
        confirmados = [c for c in confirmados if c[1] not in colisiones]
        for id_cnv, filas in colisiones.items():
            codigos = ", ".join(c[0] for c in filas)
            for codigo_cnv, _id, denominacion, _regla in filas:
                pendientes.append((
                    codigo_cnv,
                    padres[codigo_cnv][0],
                    f"colisión: el id {id_cnv} ({denominacion!r}) también lo reclaman {codigos}",
                ))
        print(f"{len(colisiones)} ids de la CNV reclamados por más de un fondo padre: a pendientes.\n")

    pendientes.sort(key=lambda p: p[0])

    total = len(padres)
    por_regla: dict[str, int] = {}
    for _codigo, _id, _denominacion, regla in confirmados:
        por_regla[regla] = por_regla.get(regla, 0) + 1

    print(
        f"{len(confirmados)}/{total} fondos padre matchean único "
        f"({len(confirmados) / total:.0%}), {len(pendientes)} a revisar a mano."
    )
    for regla, cantidad in sorted(por_regla.items()):
        print(f"    por {regla}: {cantidad}")
    print()

    if args.dry_run:
        print("--dry-run: no se escribió nada.")
        return

    hoy = date.today().isoformat()
    with open(FCI_CNV_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["codigo_cnv", "id_detalle_cnv", "denominacion_cnv", "verificado_por", "curado_en"])
        for codigo_cnv, id_cnv, denominacion, verificado_por in confirmados:
            w.writerow([codigo_cnv, id_cnv, denominacion, verificado_por, hoy])
    print(f"Escrito: {FCI_CNV_CSV}")

    with open(PENDIENTES_CSV, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["codigo_cnv", "fondo_cafci", "candidatos"])
        for codigo_cnv, nombre_cafci, detalle in pendientes:
            w.writerow([codigo_cnv, nombre_cafci, detalle])
    print(f"Escrito: {PENDIENTES_CSV} (revisión manual, no se usa en runtime)")


if __name__ == "__main__":
    main()
