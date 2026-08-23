"""Parser de la planilla diaria de CAFCI: del XLSX crudo a filas tipadas, o aborto sin filas
parciales.

Con `openpyxl`, no `pandas` — el proyecto no importa `pandas` desde `app/` (ver el docstring de
`app/ingesta/iamc/parser.py`, que es el precedente de un documento con encabezado complejo y
secciones intercaladas, aunque ahí sea PDF con `pdfplumber`; acá la mecánica de celdas es propia).

## La estructura del archivo, medida el 23/08/2026

Hoja única. Encabezado en dos filas (8 y 9): la fila 8 declara el grupo de columnas con celdas
combinadas ("Valor (mil cuotapartes)", "Variacion cuotaparte %"…) y la fila 9 el sub-encabezado
("Actual", y las **fechas base de cada variación**, que la fuente embebe ahí en vez de declararlas
aparte: "31/07/26" es la base del "mes", "30/12/25" la del "año", "31/07/25" la de "12 meses"). Los
datos empiezan en la fila 10; la primera fila de sección está en la 11.

Una fila de sección (39 en el archivo medido) tiene valor sólo en la columna A: es un título que
clasifica todo lo que sigue hasta la próxima. **No todas encabezan datos** — ver el docstring de
`secciones.py` para los divisores anidados y las notas al pie. Cualquier título fuera del catálogo
cerrado, o una fila de datos sin una sección de tipo de renta activa, aborta el parseo entero
(`PlanillaInvalida`): con la ingesta en modo wipe-and-replace, un parseo parcial significaría perder
fondos silenciosamente, y es preferible no tener la planilla del día (regla 1 del dominio).

## Lo que NO se ingiere

La columna `Reexp.Pesos` (I) se salta: es un tipo de cambio de fuente externa e inconsistente entre
filas, y la regla 3 exige derivarlo del propio universo. Los códigos internos (V, W, Y, Z, AA, AB,
AC), Decreto 596, los IDs de fondo padre/escisión (AM-AQ) y Regularización Ley 27.743 (AS) no
aportan nada que el producto use — lista de exclusión cerrada, decidida por el dueño del producto el
23/08/2026.
"""

from dataclasses import dataclass, field
from datetime import date
from io import BytesIO

import openpyxl

from app.ingesta.cafci.normalizacion import entero, fecha, numero, texto
from app.ingesta.cafci.secciones import DIVISORES, SECCIONES

FILA_INICIO_DATOS = 10
COLUMNAS_MINIMAS = 47

# Columnas 1-indexadas, tal como las lee openpyxl. Ver el docstring del módulo para el mapeo
# completo verificado contra el archivo real.
COL_FONDO = 1
COL_MONEDA = 2
COL_REGION = 3
COL_HORIZONTE = 4
COL_FECHA = 5
COL_VCP = 6
COL_VCP_ANTERIOR = 7
COL_VAR_DIARIA = 8
# COL 9 (Reexp.Pesos): no se ingiere — regla 3.
COL_VAR_MES = 10
COL_VAR_ANIO = 11
COL_VAR_12M = 12
COL_CUOTAPARTES = 13
COL_CUOTAPARTES_ANTERIOR = 14
COL_PATRIMONIO = 15
COL_PATRIMONIO_ANTERIOR = 16
COL_MARKET_SHARE = 17
COL_DEPOSITARIA = 18
COL_CODIGO_CNV = 19
COL_CALIFICACION = 20
COL_CODIGO_CAFCI = 21
COL_GERENTE = 24
COL_COMISION_INGRESO = 30
COL_HONORARIOS_ADM_SG = 31
COL_HONORARIOS_ADM_SD = 32
COL_GASTOS_ORD_GESTION = 33
COL_COMISION_RESCATE = 34
COL_COMISION_TRANSFERENCIA = 35
COL_HONORARIOS_EXITO = 36
COL_MONEDA_FONDO = 37
COL_PLAZO_LIQ = 38
COL_MINIMO_INVERSION = 44
COL_TIPO_DINERO = 46
COL_CALIFICADO = 47


class PlanillaInvalida(Exception):
    """La planilla no tiene la estructura esperada. Sin filas parciales: se aborta entera."""

    def __init__(self, motivo: str, **detalle: object):
        super().__init__(motivo)
        self.motivo = motivo
        self.detalle = detalle


@dataclass(frozen=True, slots=True)
class FilaFci:
    """Un fondo×clase, ya tipado. Los nombres de campo son 1:1 con las columnas de `public.fci`."""

    codigo_cafci: str
    fondo: str
    codigo_cnv: str | None
    seccion: str
    tipo_renta: str
    moneda: str
    region: str | None
    horizonte: str | None
    fecha_vcp: date | None
    vcp: float | None
    vcp_anterior: float | None
    var_diaria_pct: float | None
    var_mes_pct: float | None
    var_anio_pct: float | None
    var_12m_pct: float | None
    cuotapartes: float | None
    cuotapartes_anterior: float | None
    patrimonio: float | None
    patrimonio_anterior: float | None
    market_share: float | None
    gerente: str | None
    depositaria: str | None
    calificacion: str | None
    calificado: str | None
    tipo_dinero: str | None
    comision_ingreso: float | None
    honorarios_adm_sg: float | None
    honorarios_adm_sd: float | None
    gastos_ord_gestion: float | None
    comision_rescate: float | None
    comision_transferencia: float | None
    honorarios_exito: float | None
    moneda_fondo: str | None
    plazo_liq: int | None
    minimo_inversion: float | None


@dataclass(frozen=True, slots=True)
class ResultadoPlanilla:
    filas: list[FilaFci] = field(default_factory=list)
    fecha_cierre_anterior: date | None = None
    fecha_base_mes: date | None = None
    fecha_base_anio: date | None = None
    fecha_base_12m: date | None = None


def _validar_encabezado(ws: object) -> tuple[date | None, date | None, date | None, date | None]:
    """Confirma que la estructura de las filas 8-9 es la esperada y devuelve las cuatro fechas que
    la fuente embebe ahí. No compara las 47 columnas letra por letra —sería frágil ante un acento
    que cambie de codificación entre exportes— sino los anclas que, si se movieron, significan que
    el layout cambió de verdad."""
    if ws.max_column < COLUMNAS_MINIMAS:  # type: ignore[attr-defined]
        raise PlanillaInvalida(
            "la planilla trae menos columnas que las esperadas",
            columnas=ws.max_column,  # type: ignore[attr-defined]
        )

    fondo_header = texto(ws.cell(row=8, column=COL_FONDO).value)  # type: ignore[attr-defined]
    fecha_header = texto(ws.cell(row=8, column=COL_FECHA).value)  # type: ignore[attr-defined]
    actual_header = texto(ws.cell(row=9, column=COL_VCP).value)  # type: ignore[attr-defined]
    cafci_header = texto(ws.cell(row=8, column=COL_CODIGO_CAFCI).value)  # type: ignore[attr-defined]

    if fondo_header != "Fondo" or fecha_header != "Fecha" or actual_header != "Actual":
        raise PlanillaInvalida(
            "el encabezado de la planilla no tiene la forma esperada",
            fondo=fondo_header,
            fecha=fecha_header,
            actual=actual_header,
        )
    if not cafci_header or "CAFCI" not in cafci_header.upper():
        raise PlanillaInvalida(
            "la columna del Código CAFCI no está donde se esperaba", header=cafci_header
        )

    fecha_cierre_anterior = fecha(ws.cell(row=9, column=COL_VCP_ANTERIOR).value)  # type: ignore[attr-defined]
    fecha_base_mes = fecha(ws.cell(row=9, column=COL_VAR_MES).value)  # type: ignore[attr-defined]
    fecha_base_anio = fecha(ws.cell(row=9, column=COL_VAR_ANIO).value)  # type: ignore[attr-defined]
    fecha_base_12m = fecha(ws.cell(row=9, column=COL_VAR_12M).value)  # type: ignore[attr-defined]

    if fecha_cierre_anterior is None or fecha_base_mes is None or fecha_base_anio is None:
        raise PlanillaInvalida(
            "las fechas base de las variaciones no se pudieron leer del encabezado"
        )

    return fecha_cierre_anterior, fecha_base_mes, fecha_base_anio, fecha_base_12m


def _fila_de_datos(fila: int, ws: object, seccion: str, tipo_renta: str) -> FilaFci:
    def c(col: int) -> object:
        return ws.cell(row=fila, column=col).value  # type: ignore[attr-defined]

    codigo_cafci = texto(c(COL_CODIGO_CAFCI))
    if not codigo_cafci:
        raise PlanillaInvalida("fila de datos sin Código CAFCI", fila=fila, fondo=texto(c(COL_FONDO)))
    moneda = texto(c(COL_MONEDA))
    if not moneda:
        raise PlanillaInvalida("fila de datos sin Moneda", fila=fila, fondo=texto(c(COL_FONDO)))
    fondo = texto(c(COL_FONDO))
    if not fondo:
        raise PlanillaInvalida("fila de datos sin nombre de Fondo", fila=fila)

    return FilaFci(
        codigo_cafci=codigo_cafci,
        fondo=fondo,
        codigo_cnv=texto(c(COL_CODIGO_CNV)),
        seccion=seccion,
        tipo_renta=tipo_renta,
        moneda=moneda,
        region=texto(c(COL_REGION)),
        horizonte=texto(c(COL_HORIZONTE)),
        fecha_vcp=fecha(c(COL_FECHA)),
        vcp=numero(c(COL_VCP)),
        vcp_anterior=numero(c(COL_VCP_ANTERIOR)),
        var_diaria_pct=numero(c(COL_VAR_DIARIA)),
        var_mes_pct=numero(c(COL_VAR_MES)),
        var_anio_pct=numero(c(COL_VAR_ANIO)),
        var_12m_pct=numero(c(COL_VAR_12M)),
        cuotapartes=numero(c(COL_CUOTAPARTES)),
        cuotapartes_anterior=numero(c(COL_CUOTAPARTES_ANTERIOR)),
        patrimonio=numero(c(COL_PATRIMONIO)),
        patrimonio_anterior=numero(c(COL_PATRIMONIO_ANTERIOR)),
        market_share=numero(c(COL_MARKET_SHARE)),
        gerente=texto(c(COL_GERENTE)),
        depositaria=texto(c(COL_DEPOSITARIA)),
        calificacion=texto(c(COL_CALIFICACION)),
        calificado=texto(c(COL_CALIFICADO)),
        tipo_dinero=texto(c(COL_TIPO_DINERO)),
        comision_ingreso=numero(c(COL_COMISION_INGRESO)),
        honorarios_adm_sg=numero(c(COL_HONORARIOS_ADM_SG)),
        honorarios_adm_sd=numero(c(COL_HONORARIOS_ADM_SD)),
        gastos_ord_gestion=numero(c(COL_GASTOS_ORD_GESTION)),
        comision_rescate=numero(c(COL_COMISION_RESCATE)),
        comision_transferencia=numero(c(COL_COMISION_TRANSFERENCIA)),
        honorarios_exito=numero(c(COL_HONORARIOS_EXITO)),
        moneda_fondo=texto(c(COL_MONEDA_FONDO)),
        plazo_liq=entero(c(COL_PLAZO_LIQ)),
        minimo_inversion=numero(c(COL_MINIMO_INVERSION)),
    )


def parsear_planilla(contenido: bytes) -> ResultadoPlanilla:
    """El XLSX crudo, convertido en filas tipadas. Lanza `PlanillaInvalida` ante cualquier
    estructura que el catálogo cerrado de `secciones.py` no reconoce — nunca devuelve una lista
    parcial."""
    # Sin `read_only=True`: en modo read-only, `Worksheet.cell(row, column)` accede a la celda
    # re-posicionando el cursor de lectura del XML en cada llamada, y sobre ~4.300 filas × 47
    # columnas eso tarda minutos en vez de segundos (medido). Con el libro cargado entero en
    # memoria, `.cell()` es O(1) — el archivo pesa ~950 KB, trivial para cargar completo.
    try:
        libro = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    except Exception as exc:  # noqa: BLE001 — cualquier XLSX ilegible es una planilla inválida
        raise PlanillaInvalida(f"el archivo no es un XLSX legible: {exc}") from exc

    try:
        ws = libro.worksheets[0]
    except IndexError as exc:
        raise PlanillaInvalida("el libro no tiene ninguna hoja") from exc

    fecha_cierre_anterior, fecha_base_mes, fecha_base_anio, fecha_base_12m = _validar_encabezado(ws)

    filas: list[FilaFci] = []
    vistos: set[str] = set()
    seccion_actual: str | None = None
    tipo_renta_actual: str | None = None

    for fila in range(FILA_INICIO_DATOS, ws.max_row + 1):  # type: ignore[attr-defined]
        col_a = ws.cell(row=fila, column=COL_FONDO).value  # type: ignore[attr-defined]
        if col_a is None:
            continue
        col_b = ws.cell(row=fila, column=COL_MONEDA).value  # type: ignore[attr-defined]

        if col_b is None:
            titulo = texto(col_a)
            if titulo in SECCIONES:
                seccion_actual = titulo
                tipo_renta_actual = SECCIONES[titulo]
            elif titulo in DIVISORES:
                seccion_actual = titulo
                tipo_renta_actual = None
            else:
                raise PlanillaInvalida(f"sección desconocida: {titulo!r}", fila=fila)
            continue

        if tipo_renta_actual is None or seccion_actual is None:
            raise PlanillaInvalida(
                "fila de datos sin una sección de tipo de renta reconocida activa",
                fila=fila,
                fondo=texto(col_a),
            )

        fila_fci = _fila_de_datos(fila, ws, seccion_actual, tipo_renta_actual)
        if fila_fci.codigo_cafci in vistos:
            raise PlanillaInvalida(
                f"Código CAFCI duplicado: {fila_fci.codigo_cafci}", fila=fila
            )
        vistos.add(fila_fci.codigo_cafci)
        filas.append(fila_fci)

    if not filas:
        raise PlanillaInvalida("la planilla no trajo ninguna fila de datos")

    return ResultadoPlanilla(
        filas=filas,
        fecha_cierre_anterior=fecha_cierre_anterior,
        fecha_base_mes=fecha_base_mes,
        fecha_base_anio=fecha_base_anio,
        fecha_base_12m=fecha_base_12m,
    )
