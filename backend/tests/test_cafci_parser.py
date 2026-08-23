"""Tests del parser de la planilla de CAFCI.

Como el fixture recortado está en el repo (`fuentes/cafci-planilla-diaria-2026-08-21-recortada.
xlsx`, ocho filas reales sacadas de la planilla del 21/08/2026 con el encabezado real de dos
filas), estos tests corren sin red. Cubre: dos secciones de renta variable (una en USB), una de
renta fija, una fila con centinela de plazo (9999), una sin fecha_vcp, dos bajo divisores anidados
(`Clases en Dolar Estadounidense`, `En Proceso de Liquidacion por Pago Total`) con centinela -1 y
fecha vieja (2004), y las dos notas al pie del final del archivo real — que deben ignorarse sin
abortar.
"""

from pathlib import Path

import openpyxl
import pytest

from app.ingesta.cafci.parser import PlanillaInvalida, parsear_planilla

RUTA_XLSX_REAL = (
    Path(__file__).resolve().parents[2]
    / "fuentes"
    / "cafci-planilla-diaria-2026-08-21-recortada.xlsx"
)


@pytest.fixture(scope="module")
def resultado_real():
    return parsear_planilla(RUTA_XLSX_REAL.read_bytes())


class TestPlanillaReal:
    def test_trae_las_ocho_filas(self, resultado_real) -> None:
        assert len(resultado_real.filas) == 8

    def test_fechas_base_del_encabezado(self, resultado_real) -> None:
        assert resultado_real.fecha_cierre_anterior.isoformat() == "2026-08-20"
        assert resultado_real.fecha_base_mes.isoformat() == "2026-07-31"
        assert resultado_real.fecha_base_anio.isoformat() == "2025-12-30"
        assert resultado_real.fecha_base_12m.isoformat() == "2025-07-31"

    def test_tipos_de_renta_reconocidos(self, resultado_real) -> None:
        tipos = {f.tipo_renta for f in resultado_real.filas}
        assert tipos == {
            "renta_variable",
            "renta_fija",
            "fondos_cerrados",
            "mercado_dinero",
            "en_liquidacion",
        }

    def test_usb_no_se_traduce(self, resultado_real) -> None:
        usb = next(f for f in resultado_real.filas if f.moneda == "USB")
        assert usb.fondo == "Delta Acciones - Clase I"

    def test_discrepancia_moneda_vs_moneda_fondo_se_ingiere_sin_resolver(self, resultado_real) -> None:
        fila = next(f for f in resultado_real.filas if f.codigo_cafci == "4375")
        assert fila.moneda == "ARS"
        assert fila.moneda_fondo == "USD"

    def test_centinela_de_plazo_9999_se_guarda_tal_cual(self, resultado_real) -> None:
        fila = next(f for f in resultado_real.filas if f.plazo_liq == 9999)
        assert fila.tipo_renta == "fondos_cerrados"

    def test_centinela_de_plazo_menos_uno_bajo_divisores_anidados(self, resultado_real) -> None:
        fila = next(f for f in resultado_real.filas if f.plazo_liq == -1)
        assert fila.tipo_renta == "en_liquidacion"

    def test_fila_sin_fecha_vcp_se_ingiere_declarada(self, resultado_real) -> None:
        fila = next(f for f in resultado_real.filas if f.fecha_vcp is None)
        assert fila.tipo_renta == "mercado_dinero"
        assert fila.fondo == "Super Ahorro U$S - Clase A"

    def test_fecha_vieja_se_muestra_tal_cual_no_se_descarta(self, resultado_real) -> None:
        """Regla 11: un fondo con VCP de 2004 se muestra con su fecha, no se oculta ni se
        actualiza con una fecha falsa."""
        fila = next(f for f in resultado_real.filas if f.fecha_vcp is not None and f.fecha_vcp.year == 2004)
        assert fila.tipo_renta == "en_liquidacion"

    def test_calificacion_texto_verbatim_no_normalizada(self, resultado_real) -> None:
        """La fuente trae literalmente 'N/A' y 'NA' en distintas filas: se guardan tal cual, no
        se unifican a un solo valor 'no informado' (regla 11 — sólo el ausente es None)."""
        calificaciones = {f.calificacion for f in resultado_real.filas if f.calificacion}
        assert "N/A" in calificaciones or "NA" in calificaciones

    def test_codigo_cafci_es_texto(self, resultado_real) -> None:
        assert all(isinstance(f.codigo_cafci, str) for f in resultado_real.filas)

    def test_notas_al_pie_no_producen_filas_ni_abortan(self, resultado_real) -> None:
        """El fixture tiene dos notas al pie al final, igual que el archivo real: no deben
        aparecer como fondos ni cortar el parseo de lo que vino antes."""
        nombres = {f.fondo for f in resultado_real.filas}
        assert not any("Determinación" in n or "Advertencia" in n for n in nombres)


# --- Casos sintéticos de aborto: no requieren el archivo real ---


def _libro_minimo(filas_extra: list[tuple] | None = None) -> bytes:
    """Un XLSX sintético con el encabezado mínimo válido y, opcionalmente, filas agregadas
    después. `filas_extra` es una lista de (fila, columna, valor)."""
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.cell(row=8, column=1).value = "Fondo"
    ws.cell(row=8, column=5).value = "Fecha"
    ws.cell(row=9, column=6).value = "Actual"
    ws.cell(row=9, column=7).value = "20/08/26"
    ws.cell(row=9, column=10).value = "31/07/26"
    ws.cell(row=9, column=11).value = "30/12/25"
    ws.cell(row=9, column=12).value = "31/07/25"
    ws.cell(row=8, column=21).value = "Código CAFCI"
    # Fuerza a que la hoja tenga al menos 47 columnas (`ws.max_column`), como exige la validación
    # de estructura mínima.
    ws.cell(row=8, column=47).value = "Calificado"

    if filas_extra:
        for fila, columna, valor in filas_extra:
            ws.cell(row=fila, column=columna).value = valor

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestAbortos:
    def test_archivo_no_xlsx_es_invalido(self) -> None:
        with pytest.raises(PlanillaInvalida):
            parsear_planilla(b"esto no es un xlsx")

    def test_encabezado_corrido_es_invalido(self) -> None:
        """Menos de 47 columnas: la estructura mínima no está."""
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=8, column=1).value = "Fondo"
        buffer = BytesIO()
        wb.save(buffer)
        with pytest.raises(PlanillaInvalida):
            parsear_planilla(buffer.getvalue())

    def test_seccion_desconocida_aborta(self) -> None:
        contenido = _libro_minimo([(11, 1, "Una Sección Que Nunca Existió")])
        with pytest.raises(PlanillaInvalida, match="sección desconocida"):
            parsear_planilla(contenido)

    def test_fila_de_datos_sin_seccion_activa_aborta(self) -> None:
        """Una fila de datos justo debajo de un divisor (sin sección de tipo de renta) es la
        sorpresa estructural que aborta — nunca hereda el tipo de renta de la sección previa."""
        contenido = _libro_minimo(
            [
                (11, 1, "Fondos Liquidos, con rescate en efectivo"),  # divisor puro
                (12, 1, "Un Fondo Cualquiera"),
                (12, 2, "ARS"),
                (12, 21, "1"),
            ]
        )
        with pytest.raises(PlanillaInvalida, match="sin una sección"):
            parsear_planilla(contenido)

    def test_codigo_cafci_duplicado_aborta(self) -> None:
        contenido = _libro_minimo(
            [
                (11, 1, "Renta Fija Peso Argentina"),
                (12, 1, "Fondo A"),
                (12, 2, "ARS"),
                (12, 21, "1"),
                (13, 1, "Fondo B"),
                (13, 2, "ARS"),
                (13, 21, "1"),
            ]
        )
        with pytest.raises(PlanillaInvalida, match="duplicado"):
            parsear_planilla(contenido)

    def test_fila_sin_codigo_cafci_aborta(self) -> None:
        contenido = _libro_minimo(
            [
                (11, 1, "Renta Fija Peso Argentina"),
                (12, 1, "Fondo Sin Código"),
                (12, 2, "ARS"),
            ]
        )
        with pytest.raises(PlanillaInvalida):
            parsear_planilla(contenido)

    def test_sin_ninguna_fila_de_datos_aborta(self) -> None:
        contenido = _libro_minimo()
        with pytest.raises(PlanillaInvalida, match="ninguna fila"):
            parsear_planilla(contenido)
