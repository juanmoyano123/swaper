"""La ingesta de CAFCI dentro de la corrida matinal, y el job manual `correr_fci` — F-057.

Reusa el andamiaje de `test_jobs_corridas.py` (BYMA y data912 montados con respx, `_no_dormir`)
para que la matinal completa siga corriendo; lo nuevo acá es el tramo de CAFCI.
"""

import httpx
import pytest
import respx

from app.core.config import get_settings
from app.jobs.corridas import correr_fci, corrida_matinal
from tests.conftest import FakeConexionEscritura
from tests.test_jobs_corridas import (
    BYMA_URL,
    DATA912_URL,
    _montar_byma,
    _montar_data912,
    _no_dormir,
)

CAFCI_URL = "https://cafci-test.local/pb_get"


class _FakeConexionConRegistro(FakeConexionEscritura):
    async def fetchrow(self, query: str, *args):
        if "INSERT INTO public.corridas_ingesta" in query:
            self._registrar(query)
            tipo, iniciado_en, finalizado_en, duracion_ms, filas, alertas, estado = args
            return {
                "id": 1,
                "tipo": tipo,
                "iniciado_en": iniciado_en,
                "finalizado_en": finalizado_en,
                "duracion_ms": duracion_ms,
                "filas_por_fuente": filas,
                "alertas": alertas,
                "estado": estado,
            }
        return await super().fetchrow(query, *args)


@pytest.fixture
def settings_de_prueba():
    return get_settings().model_copy(
        update={
            "byma_base_url": BYMA_URL,
            "data912_base_url": DATA912_URL,
            "cafci_url": CAFCI_URL,
        }
    )


def _xlsx_minimo() -> bytes:
    """Un XLSX válido de una sola fila, suficiente para que `parsear_planilla` lo acepte."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws.cell(row=8, column=1).value = "Fondo"
    ws.cell(row=8, column=5).value = "Fecha"
    ws.cell(row=9, column=6).value = "Actual"
    ws.cell(row=9, column=7).value = "20/08/26"
    ws.cell(row=9, column=10).value = "31/07/26"
    ws.cell(row=9, column=11).value = "30/12/25"
    ws.cell(row=9, column=12).value = "31/07/25"
    ws.cell(row=8, column=21).value = "Código CAFCI"
    ws.cell(row=8, column=47).value = "Calificado"
    ws.cell(row=11, column=1).value = "Renta Fija Peso Argentina"
    ws.cell(row=12, column=1).value = "Fondo de Prueba"
    ws.cell(row=12, column=2).value = "ARS"
    ws.cell(row=12, column=21).value = "1"
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _montar_cafci_ok() -> None:
    respx.get(CAFCI_URL).mock(
        return_value=httpx.Response(
            200,
            content=_xlsx_minimo(),
            headers={"content-disposition": 'attachment; filename="20260821_Planilla_Diaria_A.xlsx"'},
        )
    )


def _montar_cafci_caida() -> None:
    respx.get(CAFCI_URL).mock(return_value=httpx.Response(500))


async def test_matinal_con_cafci_deshabilitado_no_pide_nada(settings_de_prueba) -> None:
    """Default `cafci_habilitado=False`: la matinal sigue corriendo, sin tocar la fuente."""
    conn = _FakeConexionConRegistro()
    with respx.mock:
        _montar_byma({})
        _montar_data912()
        corrida = await corrida_matinal(conn, settings_de_prueba, dormir=_no_dormir)

    assert corrida["filas_por_fuente"]["cafci"] == 0
    assert not conn.escribio_en("fci")


async def test_matinal_con_cafci_habilitado_registra_las_filas(settings_de_prueba) -> None:
    settings = settings_de_prueba.model_copy(update={"cafci_habilitado": True})
    conn = _FakeConexionConRegistro()
    with respx.mock:
        _montar_byma({})
        _montar_data912()
        _montar_cafci_ok()
        corrida = await corrida_matinal(conn, settings, dormir=_no_dormir)

    assert corrida["filas_por_fuente"]["cafci"] == 1
    assert conn.escribio_en("fci")


async def test_cafci_caida_no_aborta_la_matinal(settings_de_prueba) -> None:
    """BYMA y data912 siguen entregando lo suyo aunque CAFCI esté abajo — mismo contrato que IAMC:
    una fuente caída no tira abajo a las demás."""
    settings = settings_de_prueba.model_copy(update={"cafci_habilitado": True})
    conn = _FakeConexionConRegistro()
    with respx.mock:
        _montar_byma({})
        _montar_data912()
        _montar_cafci_caida()
        corrida = await corrida_matinal(conn, settings, dormir=_no_dormir)

    assert corrida["filas_por_fuente"]["cafci"] == 0
    assert "byma" in corrida["filas_por_fuente"]
    assert conn.escribio_en("instrumentos")


async def test_correr_fci_registra_su_propio_tipo_de_corrida(settings_de_prueba) -> None:
    settings = settings_de_prueba.model_copy(update={"cafci_habilitado": True})
    conn = _FakeConexionConRegistro()
    with respx.mock:
        _montar_cafci_ok()
        corrida = await correr_fci(conn, settings, dormir=_no_dormir)

    assert corrida["tipo"] == "fci"
    assert corrida["filas_por_fuente"] == {"cafci": 1}
    assert corrida["estado"] == "completa"


async def test_correr_fci_con_flag_apagado_registra_fallida(settings_de_prueba) -> None:
    """0 filas, sin error: `_estado_de` la marca `fallida` — el asesor que dispara el job a mano
    tiene que ver que el flag está apagado, no una corrida vacía disfrazada de éxito."""
    conn = _FakeConexionConRegistro()
    corrida = await correr_fci(conn, settings_de_prueba, dormir=_no_dormir)

    assert corrida["estado"] == "fallida"
    assert corrida["filas_por_fuente"] == {"cafci": 0}
